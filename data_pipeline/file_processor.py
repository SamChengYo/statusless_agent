import os, shutil
import re
import pandas as pd
from PyPDF2 import PdfReader, PdfWriter
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence.models import DocumentContentFormat
from docx import Document
from PIL import Image
from openpyxl import load_workbook
from pathlib import Path
import fitz  # PyMuPDF
import io
from data_pipeline.storage_operate import StorageOperate

# ========= 通用工具 =========
def _save_image(pix, path):
    img_bytes = pix.tobytes("png")
    Image.open(io.BytesIO(img_bytes)).save(path)

def _is_adjacent_or_overlapping(r1, r2, padding=5):
    return not (
        r1.x1 + padding < r2.x0 or
        r1.x0 - padding > r2.x1 or
        r1.y1 + padding < r2.y0 or
        r1.y0 - padding > r2.y1
    )

def _group_adjacent_images(img_rects_with_paths, padding=5):
    groups, used = [], set()
    for i, (rect_i, xref_i, path_i) in enumerate(img_rects_with_paths):
        if i in used:
            continue
        group = [(rect_i, xref_i, path_i)]
        used.add(i)
        q = [rect_i]
        while q:
            cur = q.pop()
            for j, (rect_j, xref_j, path_j) in enumerate(img_rects_with_paths):
                if j in used:
                    continue
                if _is_adjacent_or_overlapping(cur, rect_j, padding):
                    group.append((rect_j, xref_j, path_j))
                    used.add(j)
                    q.append(rect_j)
        groups.append(group)
    return groups

def _get_group_bbox(rects, txt_h=12):
    x0 = min(r.x0 for r in rects)
    y0 = min(r.y0 for r in rects)
    x1 = max(r.x1 for r in rects)
    y1 = max(r.y1 for r in rects)
    group_height = y1 - y0
    y0_new = y0 + group_height / 2
    y1_new = y1 + group_height / 2
    return fitz.Rect(x0, y0_new, x1, y1_new)

def _merge_images(image_paths, output_path, storage_instance: StorageOperate, direction="vertical"):
    imgs = [Image.open(p) for p in image_paths]
    if direction == "vertical":
        w, h = max(i.width for i in imgs), sum(i.height for i in imgs)
        merged = Image.new("RGB", (w, h))
        y = 0
        for i in imgs:
            merged.paste(i, (0, y)); y += i.height
    else:
        h, w = max(i.height for i in imgs), sum(i.width for i in imgs)
        merged = Image.new("RGB", (w, h))
        x = 0
        for i in imgs:
            merged.paste(i, (x, 0)); x += i.width
    os.makedirs(Path(output_path).parent, exist_ok=True)
    merged.save(output_path, "JPEG")
    for i in imgs: i.close()
    remote_key = f"images/{Path(output_path).name}"
    url = storage_instance.upload_file_return_url(output_path, remote_key)
    os.remove(output_path)
    return url

def _split_to_single_pdf(page, pdf_name, temp_dir="temp"):
    os.makedirs(temp_dir, exist_ok=True)
    single_page_pdf_path = os.path.join(temp_dir, pdf_name)
    writer = PdfWriter()
    writer.add_page(page)
    with open(single_page_pdf_path, "wb") as f:
        writer.write(f)

# === PDF 圖片處理 ===
def _preprocess_pdf_image(input_pdf: str, output_pdf: str, img_dir, storage_instance) -> list:
    doc = fitz.open(os.path.join("temp", input_pdf))
    image_paths = []
    for page_idx in range(len(doc)):
        page = doc[page_idx]
        img_rects, tmp_paths = [], []
        for img_idx, img in enumerate(page.get_images(full=True)):
            xref = img[0]
            rects = page.get_image_rects(xref)
            if not rects: continue
            pix = fitz.Pixmap(doc, xref)
            tmp_path = os.path.join("temp", f"p{page_idx}_img{img_idx}.jpg")
            _save_image(pix, tmp_path)
            img_rects.append((rects[0], xref))
            tmp_paths.append(tmp_path)
        if not img_rects: continue
        rects_paths = [(r, x, p) for (r, x), p in zip(img_rects, tmp_paths)]
        groups = _group_adjacent_images(rects_paths, padding=5)
        for g_idx, grp in enumerate(groups):
            g_paths = [p for _, _, p in grp]
            if len(grp) == 1:
                mdir = "vertical"
            else:
                rs = sorted([r for r, _, _ in grp], key=lambda r: (r.y0, r.x0))
                dy = abs(rs[1].y0 - rs[0].y1)
                dx = abs(rs[1].x0 - rs[0].x1)
                mdir = "vertical" if dy < dx else "horizontal"
            local_name = f"{Path(input_pdf).stem}_g{g_idx}.jpg"
            local_path = Path(img_dir) / local_name
            url = _merge_images(
                [str(p) for p in g_paths],
                str(local_path),
                storage_instance,
                direction=mdir,
            )
            image_paths.append(url)
            for rect, xref, _ in grp:
                try:
                    doc._deleteObject(xref)
                except Exception:
                    pass
            md = f"![{local_name}]({url})"
            bbox = _get_group_bbox([r for r, _, _ in grp])
            page.insert_textbox(bbox, md, fontsize=10, fontname="helv")
        for p in tmp_paths:
            Path(p).unlink(missing_ok=True)
    doc.save(output_pdf, deflate=True, garbage=4)
    doc.close()
    shutil.rmtree("temp", ignore_errors=True)
    return image_paths

class DocumentProcessor:
    def __init__(self, storage_instance: StorageOperate, split_dir, md_dir, img_dir):
        self.azure_endpoint = os.getenv("AZURE_AI_INTELLIGENCE_ENDPOINT")
        self.azure_api_key = os.getenv("AZURE_AI_INTELLIGENCE_API_KEY")
        self.SPLIT_DIR = split_dir
        self.MD_DIR = md_dir
        self.IMG_DIR = img_dir
        self.storage_instance = storage_instance
        for _dir in [self.SPLIT_DIR, self.MD_DIR, self.IMG_DIR]:
            os.makedirs(_dir, exist_ok=True)
    def azure_doc(self, file_path):
        client = DocumentIntelligenceClient(endpoint=self.azure_endpoint, credential=AzureKeyCredential(self.azure_api_key))
        with open(file_path, "rb") as f:
            poller = client.begin_analyze_document(
                model_id="prebuilt-layout",
                body=f,
                output_content_format=DocumentContentFormat.MARKDOWN
            )
        documents = poller.result()
        return documents["content"]
    def process_pdf_file(self, pdf_file_path):
        filename = os.path.basename(pdf_file_path)
        reader = PdfReader(pdf_file_path)
        md_result = []
        for page_number, page in enumerate(reader.pages, start=1):
            single_page_name = f"{os.path.splitext(filename)[0]}_page_{page_number}.pdf"
            _split_to_single_pdf(page, single_page_name)
            parsed_pdf_output_path = os.path.join(self.SPLIT_DIR, single_page_name)
            image_paths = _preprocess_pdf_image(single_page_name, parsed_pdf_output_path, self.IMG_DIR, self.storage_instance)
            md_content = self.azure_doc(parsed_pdf_output_path)
            md_markdown = {"page": page_number, "markdown": md_content.replace("!\[", "!["), "image_urls": image_paths}
            md_result.append({"md_result": md_markdown})
        return md_result
    def process_image_file(self, file_path):
        md = self.azure_doc(file_path)
        return [{"md_result": {"markdown": md}}]
    def process_csv_file(self, csv_path):
        df = pd.read_csv(csv_path)
        return [{"md_result": {"markdown": df.to_markdown(index=False)}}]
    def process_docx_file(self, docx_path):
        doc = Document(docx_path)
        paras = [p.text for p in doc.paragraphs if p.text.strip()]
        return [{"md_result": {"markdown": "\n\n".join(paras)}}]
    def process_xlsx_file(self, xlsx_path):
        wb = load_workbook(xlsx_path, data_only=True)
        results = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            df = pd.DataFrame(ws.values)
            if df.dropna(how='all').empty: continue
            df.columns = df.iloc[0]
            df = df[1:]
            md = f"# 工作表：{sheet}\n" + df.to_markdown(index=False)
            results.append({"md_result": {"page": sheet, "markdown": md}})
        return results
    def process_txt_file(self, txt_path):
        content = Path(txt_path).read_text(encoding="utf-8")
        return [{"md_result": {"markdown": content}}]
    def save_md_to_blob(self, md_content, filename):
        md_path = Path(self.MD_DIR)/f"{filename}.md"
        md_path.write_text(md_content, encoding="utf-8")
        blob_key = f"tmp_md/{filename}.md"
        # 假設 storage_instance 同樣支援 upload_fileobj_return_url
        with open(md_path, "rb") as f:
            url = self.storage_instance.upload_fileobj_return_url(f, blob_key)
        return url
    def process_file_streaming(self, file_path):
        ext = Path(file_path).suffix.lower()
        name = Path(file_path).stem

        if ext == ".pdf":
            results = self.process_pdf_file(file_path)
        elif ext in [".jpg", ".jpeg", ".png"]:
            results = self.process_image_file(file_path)
        elif ext == ".txt":
            results = self.process_txt_file(file_path)
        elif ext == ".csv":
            results = self.process_csv_file(file_path)
        elif ext in [".xlsx", ".xls"]:
            results = self.process_xlsx_file(file_path)
        elif ext == ".docx":
            results = self.process_docx_file(file_path)
        else:
            yield {"error": "Unsupported file type"}
            return

        for i, r in enumerate(results, 1):
            try:
                md = r["md_result"]["markdown"]
                url = self.save_md_to_blob(md, f"{name}_part_{i}")
                yield {
                    "status": "success",
                    "part": i,
                    "message": f"Page {i} processed",
                    "url": url
                }
            except Exception as e:
                yield {
                    "status": "error",
                    "part": i,
                    "message": f"Error processing page {i}: {str(e)}",
                    "url": None
                }
